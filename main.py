from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook, Workbook
import re
import os

#recebedor de arquivo!
    #Verifica peso do arquivo! ‚ùå
    #verifica se o que o cara digitou √© vazio‚ùå 
    #testa se √© um arquivo xml! ‚úîÔ∏è
#verificador de link
    #verificar se tem a estrutura de um link e o adiciona na lista! ‚ùå
    #verificar espa√ßos em brancos no xml e o add na lista! ‚úîÔ∏è
    #reformular os c√≥digos e tirar esse mundo de lista! ‚ùå
#distribui os links em htmls
    #cria√ß√£o da pasta para inser√ß√£o dos htmls!‚ùå
    #cria√ß√£o arquivo htmls com o resultado do get!‚ùå
"""
Pessoa introduz um xml, ele pegar√° todos os links(tirando os espa√ßos vazios , descartando os digitados errados, linhas em brancos, coisas que
forem links!), dar√° um get e salvar√° o retorno em um arquivo a parte!

"""


class Logic:
    def __init__(self) -> None:
        self.links_raw = []  # lista para armazenar as tags diretamente do xml
        
        self.links_validated = []  # lista com os links v√°lidos!
        self.links_error = []  # lista com os links err√¥neos
        self.supose_path_xml = None #path do xml!
        
    def create_path_html(self,path):
        pasta_name = 'Resultado_xml!'
        path_complete = os.path.join(path,pasta_name)
        if not os.path.exists(path_complete):
            os.makedirs(path_complete)
            print('Pasta criada com sucesso!')
            
    def get_paths_xml_lib(self):
        self.supose_path_xml = input('Digite o diret√≥rio do arquivo XML!')
        if self.supose_path_xml.endswith('.xml'):
            self.supose_path_lib = input('Digite agora onde deseja salvar a pasta de retorno do xml! ')
            self.create_path_html(self.supose_path_lib)
            return self.supose_path_xml
            
            
    def path_tester_xml(self):#testa se √© um xml o que a pessoa est√° introduzindo!
        self.supose_path_xml = input('Digite o diret√≥rio do arquivo XML!')
        
        # self.create_path_html(self.supose_path_xml)
        if '.xml'in self.supose_path_xml: #.endswith('.xml')
            return True,self.supose_path_xml
        return False,None
    
    def empty_input():
        answer = input('>')
        if not answer:
            return True

            
        
        
    def get_line_from_xml(self):  # pega linha por linha do xml e armazena na lista 'links_raw'
        with open(self.supose_path_xml, "r") as file:
            line = file.read()
        content = BeautifulSoup(line, "xml")

        for item in content.find_all("site"):  # salva apenas o link na lista
            
            if item.text.strip() == '':
                continue
            self.links_raw.append(item.text)

        return self.links_raw  # retornando em lista com os valores
    
    def check_if_its_url(self): #verifica se o link funciona , dividindo entre os funcionais e os n√£o funcionais!
        pattern_url = r'https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+(\.[a-zA-Z]{2,})(?:/.*)?'
            
        for links in self.links_raw:
                           
            if re.match(pattern_url,links):
                self.links_validated.append(links)
            
            else:
                       
                self.links_error.append(links)

    def check_if_url_works_xml(self):  # testa as urls e as separa se funfa ou n√£o!
        for urls in self.links_raw:
            try:
                maybe_url = requests.get(urls)
                if maybe_url.status_code == 200:
                    self.links_validated.append(urls)

            except:
                self.links_error.append(urls)

        return self.links_validated, self.links_error


        
        
        
        ...
        
    def create_htmls():
        ...
    
    def add_link_xlsx(self):  # passa os links para a planilha! #remover isso aqui!!üî¥
        try:
            # carrega a planilha caso ela n√£o exista
            work_book = load_workbook("links.xlsx")
            workbook_active = work_book.active
        except:
            # cria a planilha!
            work_book = Workbook()
            workbook_active = work_book.active
        for link in self.links_validated:  # adiciona os links v√°lidos
            workbook_active.append([link, "Link v√°lido!"])
        for link in self.links_error:  # adiciona os links n√£o v√°lidos!
            workbook_active.append([link, "Link inv√°lido!"])
        work_book.save("links.xlsx")
    
  

a = Logic()

a.get_paths()
print(a.get_line_from_xml())
a.check_if_its_url()
print(a.links_validated,'--',a.links_error)


